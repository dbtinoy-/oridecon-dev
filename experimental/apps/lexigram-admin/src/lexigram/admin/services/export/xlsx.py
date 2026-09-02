"""Shared Excel (.xlsx) row encoder for admin exports (R29).

One xlsx writer serves both export paths:

* the background job flow (``ExcelExportBackend.generate_file``), and
* the direct-download bulk/toolbar exports in both stacks.

openpyxl is an optional dependency (same guarded-import pattern as the
R26 xlsx importer); :func:`encode_rows_as_xlsx` raises ``ImportError``
when it is absent so HTTP callers can answer with a clear status instead
of a 500.

Cell handling:

* every value passes through :func:`sanitize_cell_value` — Excel
  evaluates leading ``=+-@`` formulas exactly like CSV importers do;
* values openpyxl cannot store natively (dict/list/set/tuple/bytes/
  arbitrary objects) are stringified by :func:`coerce_cell_value` —
  previously a JSON column crashed the Excel backend with
  ``ValueError: Cannot convert … to Excel``.
"""

from __future__ import annotations

from datetime import date, datetime, time
from decimal import Decimal
import io
from typing import Any

from lexigram.admin.services.export.sanitize import sanitize_cell_value

try:
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Font,
        PatternFill,
    )

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover — exercised via monkeypatch in tests
    HAS_OPENPYXL = False

#: Types openpyxl stores natively; everything else is stringified.
_NATIVE_CELL_TYPES = (str, int, float, bool, datetime, date, time, Decimal)

#: Width autosize: rows sampled and per-column character cap.
_WIDTH_SAMPLE_ROWS = 100
_MAX_COLUMN_WIDTH = 50

#: Header fill (parity with the original ExcelExportBackend styling).
_HEADER_FILL_COLOR = "FFE6E6FA"

OPENPYXL_MISSING_MESSAGE = (
    "Excel export requires the optional 'openpyxl' dependency (pip install openpyxl)."
)


def coerce_cell_value(value: Any) -> Any:
    """Return a value openpyxl can store in a cell.

    Native scalar types (and None) pass through unchanged; bytes decode
    as UTF-8 (replacement on errors); anything else — dicts, lists, sets,
    tuples, arbitrary objects — is stringified.
    """
    if value is None or isinstance(value, _NATIVE_CELL_TYPES):
        return value
    if isinstance(value, (bytes, bytearray)):
        return bytes(value).decode("utf-8", errors="replace")
    return str(value)


def encode_rows_as_xlsx(
    rows: list[dict[str, Any]],
    fieldnames: list[str] | None = None,
    *,
    sheet_title: str = "Export",
) -> bytes:
    """Encode dict rows as a styled single-sheet ``.xlsx`` workbook.

    Args:
        rows: Export rows; keys become columns.
        fieldnames: Explicit column order. Defaults to first-seen key
            order across all rows (union — later rows may add columns).
        sheet_title: Worksheet title.

    Returns:
        The workbook serialized to bytes.

    Raises:
        ImportError: When openpyxl is not installed.
    """
    if not HAS_OPENPYXL:
        raise ImportError(OPENPYXL_MISSING_MESSAGE)

    if fieldnames is None:
        fieldnames = []
        for row in rows:
            for key in row:
                if key not in fieldnames:
                    fieldnames.append(str(key))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title

    if fieldnames:
        header_fill = PatternFill(
            start_color=_HEADER_FILL_COLOR,
            end_color=_HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(bold=True)
        for col_num, header in enumerate(fieldnames, 1):
            cell = worksheet.cell(row=1, column=col_num, value=str(header))
            cell.font = header_font
            cell.fill = header_fill

        right_align = Alignment(horizontal="right")
        for row_num, row in enumerate(rows, 2):
            for col_num, field in enumerate(fieldnames, 1):
                value = coerce_cell_value(sanitize_cell_value(row.get(field)))
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                if isinstance(value, bool):
                    pass  # bool is an int subclass — no numeric alignment
                elif isinstance(value, (int, float, Decimal)):
                    cell.alignment = right_align
                elif isinstance(value, datetime):
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"

        # Autosize columns from a bounded sample.
        for col_num, field in enumerate(fieldnames, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            max_length = len(str(field))
            for row in rows[:_WIDTH_SAMPLE_ROWS]:
                raw = row.get(field)
                if raw is not None:
                    max_length = max(max_length, len(str(raw)))
            worksheet.column_dimensions[column_letter].width = min(
                max_length + 2,
                _MAX_COLUMN_WIDTH,
            )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


__all__ = [
    "HAS_OPENPYXL",
    "OPENPYXL_MISSING_MESSAGE",
    "XLSX_CONTENT_TYPE",
    "coerce_cell_value",
    "encode_rows_as_xlsx",
]
