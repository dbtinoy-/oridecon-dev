from __future__ import annotations

from datetime import UTC
import io
from typing import TYPE_CHECKING, Any

from lexigram.admin.services.export.sanitize import sanitize_cell_value
from lexigram.admin.services.export.service import IExportBackend

if TYPE_CHECKING:
    from lexigram.admin.services.export.scheduler import ExportJob

try:
    import openpyxl  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Font,
        PatternFill,
    )

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExportBackend(IExportBackend):
    """Backend for Excel exports."""

    async def generate_file(
        self,
        job: ExportJob,
        data: list[dict[str, Any]],
        storage: Any,
        export_dir: str,
    ) -> str:
        """Export data to Excel format."""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export")

        from datetime import datetime

        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        filename = f"{job.resource_name}_export_{timestamp}.xlsx"
        file_path = f"{export_dir}/{filename}"

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Export"

        if data:
            # Determine columns
            fieldnames = job.columns or list(data[0].keys())

            # Write headers
            for col_num, header in enumerate(fieldnames, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFE6E6FA",
                    end_color="FFE6E6FA",
                    fill_type="solid",
                )

            # Write data
            for row_num, row in enumerate(data, 2):
                for col_num, field in enumerate(fieldnames, 1):
                    value = sanitize_cell_value(row.get(field, ""))
                    cell = worksheet.cell(row=row_num, column=col_num, value=value)

                    # Basic formatting
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                    elif isinstance(value, datetime):
                        cell.number_format = "YYYY-MM-DD HH:MM:SS"

            # Auto-adjust column widths
            for col_num, field in enumerate(fieldnames, 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)
                max_length = len(field)  # Start with header length

                for row in data[:100]:  # Sample first 100 rows
                    value = str(row.get(field, ""))
                    max_length = max(max_length, len(value))

                worksheet.column_dimensions[column_letter].width = min(
                    max_length + 2,
                    50,
                )

        # Save to storage
        buffer = io.BytesIO()
        workbook.save(buffer)
        await storage.upload(
            file_path,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        return file_path
