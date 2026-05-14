"""Office and email document loaders for RAG."""

from __future__ import annotations

import asyncio
import email
import email.policy
from pathlib import Path

from lexigram.ai.rag.chunking.types import Chunk
from lexigram.ai.rag.loaders._io_utils import read_file_bytes, read_file_text
from lexigram.ai.rag.types import RAGError

# ---------------------------------------------------------------------------
# DocxLoader
# ---------------------------------------------------------------------------


class DocxLoader:
    """Load Microsoft Word (.docx) documents.

    Requires: python-docx
    Install: pip install openpyxl
    """

    async def load(self, source: str | Path) -> list[Chunk]:
        """Load a .docx file into chunks — one chunk per paragraph.

        Args:
            source: Path to the .docx file.

        Returns:
            List of chunks, one per non-empty paragraph.

        Raises:
            ImportError: If python-docx is not installed.
            RAGError: If the file cannot be read or parsed.
        """
        try:
            try:
                import docx  # type: ignore[import-not-found]
            except ImportError as e:
                msg = "DocxLoader requires 'python-docx'. Install with: pip install python-docx"
                raise ImportError(msg) from e

            path = Path(source)
            raw = await read_file_bytes(path)

            def _parse() -> list[Chunk]:
                import io as _io

                doc = docx.Document(_io.BytesIO(raw))
                result: list[Chunk] = []
                for idx, para in enumerate(doc.paragraphs):
                    text = para.text.strip()
                    if not text:
                        continue
                    result.append(
                        Chunk(
                            text=text,
                            source=str(path),
                            chunk_index=len(result),
                            metadata={
                                "source": str(path),
                                "type": "docx",
                                "paragraph_index": idx,
                                "style": para.style.name if para.style else None,
                            },
                        )
                    )
                return result

            return await asyncio.to_thread(_parse)

        except (ImportError, RAGError):
            raise
        except (OSError, FileNotFoundError) as e:
            msg = f"Failed to read DOCX file {source}: {e}"
            raise RAGError(msg) from e
        except Exception as e:
            msg = f"Unexpected error loading DOCX file {source}: {e}"
            raise RAGError(msg) from e


# ---------------------------------------------------------------------------
# ExcelLoader
# ---------------------------------------------------------------------------


class ExcelLoader:
    """Load Microsoft Excel (.xlsx / .xls) documents.

    Produces one chunk per worksheet row (or per sheet when
    ``chunk_per_sheet=True``).

    Requires: openpyxl
    Install: pip install openpyxl
    """

    def __init__(self, *, chunk_per_sheet: bool = False) -> None:
        """Initialize Excel loader.

        Args:
            chunk_per_sheet: When True, produce one chunk per sheet instead
                of one chunk per row.
        """
        self.chunk_per_sheet = chunk_per_sheet

    async def load(self, source: str | Path) -> list[Chunk]:
        """Load an Excel file.

        Args:
            source: Path to the .xlsx or .xls file.

        Returns:
            List of chunks.

        Raises:
            ImportError: If openpyxl is not installed.
            RAGError: If the file cannot be read or parsed.
        """
        try:
            try:
                import openpyxl  # type: ignore[import-untyped]
            except ImportError as e:
                msg = "ExcelLoader requires 'openpyxl'. Install with: pip install openpyxl"
                raise ImportError(msg) from e

            path = Path(source)
            raw = await read_file_bytes(path)

            def _parse() -> list[Chunk]:
                import io as _io

                wb = openpyxl.load_workbook(_io.BytesIO(raw), data_only=True)
                result: list[Chunk] = []

                for sheet in wb.worksheets:
                    if self.chunk_per_sheet:
                        rows_text = []
                        for row in sheet.iter_rows(values_only=True):
                            cells = [str(c) if c is not None else "" for c in row]
                            rows_text.append("\t".join(cells))
                        text = "\n".join(rows_text)
                        if text.strip():
                            result.append(
                                Chunk(
                                    text=text,
                                    source=str(path),
                                    chunk_index=len(result),
                                    metadata={
                                        "source": str(path),
                                        "type": "excel",
                                        "sheet": sheet.title,
                                    },
                                )
                            )
                    else:
                        headers: list[str] = []
                        for row_idx, row in enumerate(
                            sheet.iter_rows(values_only=True)
                        ):
                            if row_idx == 0:
                                headers = [
                                    str(c) if c is not None else f"col{i}"
                                    for i, c in enumerate(row)
                                ]
                                continue
                            cells = [str(c) if c is not None else "" for c in row]
                            text = " ".join(
                                f"{h}: {v}"
                                for h, v in zip(headers, cells, strict=False)
                            )
                            if text.strip():
                                result.append(
                                    Chunk(
                                        text=text,
                                        source=str(path),
                                        chunk_index=len(result),
                                        metadata={
                                            "source": str(path),
                                            "type": "excel",
                                            "sheet": sheet.title,
                                            "row": row_idx,
                                        },
                                    )
                                )
                return result

            return await asyncio.to_thread(_parse)

        except (ImportError, RAGError):
            raise
        except (OSError, FileNotFoundError) as e:
            msg = f"Failed to read Excel file {source}: {e}"
            raise RAGError(msg) from e
        except Exception as e:
            msg = f"Unexpected error loading Excel file {source}: {e}"
            raise RAGError(msg) from e


# ---------------------------------------------------------------------------
# EmailLoader
# ---------------------------------------------------------------------------


class EmailLoader:
    """Load email messages (.eml files).

    Extracts subject, from, to, date headers and the plain-text body.
    Uses only stdlib ``email`` package — no additional dependencies.
    """

    async def load(self, source: str | Path) -> list[Chunk]:
        """Load an .eml file.

        Args:
            source: Path to the .eml file.

        Returns:
            Single chunk whose text is the email body (plain-text part).

        Raises:
            RAGError: If the file cannot be read or parsed.
        """
        try:
            path = Path(source)
            raw = await read_file_text(path, encoding="utf-8")

            def _parse() -> Chunk:
                msg = email.message_from_string(raw, policy=email.policy.default)
                subject = str(msg.get("subject", ""))
                from_addr = str(msg.get("from", ""))
                to_addr = str(msg.get("to", ""))
                date = str(msg.get("date", ""))

                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            payload = part.get_payload(decode=True)
                            if isinstance(payload, bytes):
                                body = payload.decode(
                                    part.get_content_charset() or "utf-8",
                                    errors="replace",
                                )
                            break
                else:
                    payload = msg.get_payload(decode=True)
                    if isinstance(payload, bytes):
                        body = payload.decode(
                            msg.get_content_charset() or "utf-8",
                            errors="replace",
                        )

                text = f"Subject: {subject}\nFrom: {from_addr}\nTo: {to_addr}\nDate: {date}\n\n{body}"
                return Chunk(
                    text=text.strip(),
                    source=str(path),
                    chunk_index=0,
                    metadata={
                        "source": str(path),
                        "type": "email",
                        "subject": subject,
                        "from": from_addr,
                        "to": to_addr,
                        "date": date,
                    },
                )

            chunk = await asyncio.to_thread(_parse)
            return [chunk]

        except (OSError, UnicodeDecodeError) as e:
            msg = f"Failed to read email file {source}: {e}"
            raise RAGError(msg) from e
        except Exception as e:
            msg = f"Unexpected error loading email file {source}: {e}"
            raise RAGError(msg) from e


__all__ = ["DocxLoader", "EmailLoader", "ExcelLoader"]
